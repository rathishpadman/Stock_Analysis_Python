"""
Apply Conditional Formatting to Stock Analysis Excel Template
This script reads the conditional formatting rules and applies them to the NIFTY50 sheet.
"""

import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import re
import sys

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


def parse_rule(rule_str: str) -> tuple:
    """
    Parse a rule string like '< 15', '15 - 25', '> 25' into operator and value(s).
    Returns: (operator, value1, value2) where value2 is None for single-value rules.
    """
    rule_str = rule_str.strip()
    
    # Remove percentage signs and clean up
    original = rule_str
    rule_str = rule_str.replace("%", "").strip()
    
    # Special case: "< 1.0 or > 3.0" (two conditions)
    if " or " in rule_str:
        # Return as text - we'll handle this manually
        return ("text_or", original, None)
    
    # Special case: "> 0 (rising)" or "< 0 (falling)" - descriptive
    if "(" in rule_str and ")" in rule_str:
        # Extract just the numeric part
        numeric_part = rule_str.split("(")[0].strip()
        try:
            if numeric_part.startswith(">"):
                val = float(numeric_part.replace(">", "").strip())
                return (">", val, None)
            elif numeric_part.startswith("<"):
                val = float(numeric_part.replace("<", "").strip())
                return ("<", val, None)
        except ValueError:
            return ("text", original, None)
    
    # Range rule: "15 - 25"
    if " - " in rule_str:
        parts = rule_str.split(" - ")
        try:
            val1 = float(parts[0].strip())
            val2 = float(parts[1].strip())
            return ("range", val1, val2)
        except ValueError:
            return ("text", original, None)
    
    # Less than: "< 15"
    elif rule_str.startswith("<"):
        try:
            val = float(rule_str.replace("<", "").strip())
            return ("<", val, None)
        except ValueError:
            return ("text", original, None)
    
    # Greater than: "> 15"
    elif rule_str.startswith(">"):
        try:
            val = float(rule_str.replace(">", "").strip())
            return (">", val, None)
        except ValueError:
            return ("text", original, None)
    
    # Text-based rules (e.g., "Bullish", "Bearish", "rising", "falling")
    return ("text", original, None)


def apply_conditional_formatting(template_path: str, rules_csv: str, output_path: str = None):
    """
    Apply conditional formatting rules to Excel template.
    
    Args:
        template_path: Path to Excel template (.xlsx)
        rules_csv: Path to CSV file with formatting rules
        output_path: Path to save formatted Excel (defaults to template_path)
    """
    
    if output_path is None:
        output_path = template_path
    
    logger.info(f"Loading template: {template_path}")
    wb = load_workbook(template_path)
    
    # Target the NIFTY50 sheet
    sheet_name = "NIFTY50"
    if sheet_name not in wb.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        return
    
    ws = wb[sheet_name]
    logger.info(f"Applying formatting to sheet: {sheet_name}")
    
    # Load rules from CSV
    logger.info(f"Loading rules from: {rules_csv}")
    rules_df = pd.read_csv(rules_csv)
    
    # Create a mapping: Excel column name → (operator, good_val, neutral_val, poor_val)
    format_map = {}
    
    for _, row in rules_df.iterrows():
        param = row["Parameter (from file)"].strip()
        good = row["Good (Green)"].strip()
        neutral = row["Neutral (Yellow)"].strip()
        poor = row["Poor (Red)"].strip()
        
        format_map[param] = {
            "good": good,
            "neutral": neutral,
            "poor": poor,
            "category": row["Metric Category"],
            "rationale": row["Rationale & Industry Context"]
        }
    
    logger.info(f"Loaded {len(format_map)} formatting rules")
    
    # Define fill colors (Excel RGB format)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # Light red
    
    # Get header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1):
        header_row = {cell.value: cell.column for cell in row if cell.value}
        break
    
    if not header_row:
        logger.warning("Could not find header row")
        return
    
    logger.info(f"Found {len(header_row)} columns in template")
    
    # Apply conditional formatting to each column
    applied_count = 0
    for param, rules in format_map.items():
        if param not in header_row:
            logger.debug(f"Column '{param}' not found in template headers")
            continue
        
        col_num = header_row[param]
        col_letter = get_column_letter(col_num)
        
        good_rule = parse_rule(rules["good"])
        neutral_rule = parse_rule(rules["neutral"])
        poor_rule = parse_rule(rules["poor"])
        
        # Determine data range (skip header row 1)
        last_row = ws.max_row
        data_range = f"{col_letter}2:{col_letter}{last_row}"
        
        try:
            # Skip text-based rules (like "> 0 (rising)" or "< 1.0 or > 3.0")
            if good_rule[0] == "text" or good_rule[0] == "text_or":
                logger.debug(f"Skipping text-based rule for '{param}': {good_rule[1]}")
                continue
            if neutral_rule[0] == "text" or neutral_rule[0] == "text_or":
                logger.debug(f"Skipping text-based rule for '{param}': {neutral_rule[1]}")
                continue
            if poor_rule[0] == "text" or poor_rule[0] == "text_or":
                logger.debug(f"Skipping text-based rule for '{param}': {poor_rule[1]}")
                continue
            
            # Apply rules in order: Good (Green), Neutral (Yellow), Poor (Red)
            # This ensures proper precedence
            
            # Good rule (Green) - most restrictive, apply first
            if good_rule[0] == "<":
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator="lessThan", formula=[str(good_rule[1])], fill=green_fill)
                )
            elif good_rule[0] == ">":
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator="greaterThan", formula=[str(good_rule[1])], fill=green_fill)
                )
            elif good_rule[0] == "range":
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator="between", formula=[str(good_rule[1]), str(good_rule[2])], fill=green_fill)
                )
            
            # Neutral rule (Yellow) - range
            if neutral_rule[0] == "range":
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator="between", formula=[str(neutral_rule[1]), str(neutral_rule[2])], fill=yellow_fill)
                )
            
            # Poor rule (Red) - least restrictive
            if poor_rule[0] == "<":
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator="lessThan", formula=[str(poor_rule[1])], fill=red_fill)
                )
            elif poor_rule[0] == ">":
                ws.conditional_formatting.add(
                    data_range,
                    CellIsRule(operator="greaterThan", formula=[str(poor_rule[1])], fill=red_fill)
                )
            elif poor_rule[0] == "range":
                # For range, we need to apply outside logic (not between)
                pass
            
            applied_count += 1
            logger.info(f"✓ Applied formatting to '{param}' ({col_letter})")
        
        except Exception as e:
            logger.error(f"✗ Failed to apply formatting to '{param}': {e}")
    
    # Save the workbook
    logger.info(f"Saving formatted template to: {output_path}")
    wb.save(output_path)
    logger.info(f"✓ Conditional formatting applied to {applied_count} columns")
    logger.info("✓ Excel file updated successfully!")


def generate_legend_sheet(template_path: str, rules_csv: str):
    """
    Create a new sheet with the conditional formatting legend.
    """
    wb = load_workbook(template_path)
    
    # Create legend sheet
    if "Legend" in wb.sheetnames:
        del wb["Legend"]
    
    ws = wb.create_sheet("Legend")
    
    # Add title
    ws["A1"] = "Conditional Formatting Legend"
    ws["A1"].font = Font(size=14, bold=True)
    
    # Load rules
    rules_df = pd.read_csv(rules_csv)
    
    # Write header
    headers = ["Metric", "Category", "Good (Green)", "Neutral (Yellow)", "Poor (Red)", "Rationale"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Write rules
    for idx, (_, row) in enumerate(rules_df.iterrows(), 4):
        ws.cell(row=idx, column=1).value = row["Parameter (from file)"]
        ws.cell(row=idx, column=2).value = row["Metric Category"]
        ws.cell(row=idx, column=3).value = row["Good (Green)"]
        ws.cell(row=idx, column=4).value = row["Neutral (Yellow)"]
        ws.cell(row=idx, column=5).value = row["Poor (Red)"]
        ws.cell(row=idx, column=6).value = row["Rationale & Industry Context"]
    
    # Auto-adjust column widths
    for col in range(1, 7):
        max_len = 0
        column_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 50)
    
    wb.save(template_path)
    logger.info("✓ Legend sheet created")


if __name__ == "__main__":
    # Accept command-line arguments for flexibility
    if len(sys.argv) >= 3:
        template = sys.argv[1]
        rules_file = sys.argv[2]
    else:
        # Default values
        template = "Stocks_Analysis_Template_v3.xlsx"
        rules_file = "conditional_format_rules.csv"
    
    if not os.path.exists(template):
        logger.error(f"Template not found: {template}")
    elif not os.path.exists(rules_file):
        logger.error(f"Rules file not found: {rules_file}")
    else:
        apply_conditional_formatting(template, rules_file)
        generate_legend_sheet(template, rules_file)
        logger.info("\n" + "="*60)
        logger.info("✓ Conditional formatting successfully applied!")
        logger.info("="*60)
