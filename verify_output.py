#!/usr/bin/env python
"""Verify that output file has Sector P/E Median and conditional formatting on all rows."""

from openpyxl import load_workbook
import pandas as pd

# Check the newly created output file
output_path = 'OCT25/Stock_Analysis_201025_2156.xlsx'
wb = load_workbook(output_path)
sheet = wb['NIFTY50']

print('=== OUTPUT FILE VERIFICATION ===\n')

# 1. Check conditional formatting on all rows
print('1. CONDITIONAL FORMATTING:')
print(f'   Total rows: {sheet.max_row}')
print(f'   Formatting rules: {len(sheet.conditional_formatting)}')
if sheet.conditional_formatting:
    print('   Ranges applied to:')
    for i, (range_str, rules) in enumerate(sheet.conditional_formatting._cf_rules.items(), 1):
        print(f'     {i:2d}. {range_str}')

# 2. Check Sector P/E Median values
print('\n2. SECTOR P/E MEDIAN:')
df = pd.read_excel(output_path, sheet_name='NIFTY50')
if 'Sector P/E (Median)' in df.columns:
    print(f'   ✓ Column exists')
    median_values = df[['Ticker', 'Sector', 'P/E (TTM)', 'Sector P/E (Median)']].head(15)
    print('   Sample data:')
    for idx, row in median_values.iterrows():
        ticker = row['Ticker'] if 'Ticker' in row and pd.notna(row['Ticker']) else 'N/A'
        sector = row['Sector'] if pd.notna(row['Sector']) else 'N/A'
        pe = f"{row['P/E (TTM)']:.2f}" if pd.notna(row['P/E (TTM)']) else 'N/A'
        median = f"{row['Sector P/E (Median)']:.2f}" if pd.notna(row['Sector P/E (Median)']) else 'N/A'
        print(f'     {ticker:12s} {sector:20s} P/E: {pe:7s} Sector Median: {median:7s}')
else:
    print('   ✗ Column not found!')

# 3. Check non-empty Sector P/E values
print('\n3. DATA POPULATION:')
non_empty_count = df['Sector P/E (Median)'].notna().sum()
print(f'   Non-empty Sector P/E values: {non_empty_count}/{len(df)}')
